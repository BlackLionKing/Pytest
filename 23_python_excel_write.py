"""
    Yaml
        可读性高 用来表达数据序列化的格式 常常作为配置文件使用
    Json
        轻量级的数据交换语言 用来传输由属性值活着序列性的值 组成的数据对象
    Excel
        有直观的界面 出色的计算功能和图表工具 是一款制表软件

"""

# 向excel写入数据

# 导入 workbook
from openpyxl import Workbook
import datetime

# 创建对象
wb = Workbook()

# 打开Excel一个sheet
ws = wb.active
# 修改sheet的名称
ws.title = '工作表'
# 给A1单元格赋值42
ws['A1'] = 42

# 追加行数据
# 如果第1行有数据 则默认在第二行追加 1 2 3
ws.append([1, 2, 3])

# 调用datetime模块查询当前时间并赋值给a2单元格 excel将自动转换日期格式
ws['A2'] = datetime.datetime.now()

# 创建新工作表my_sheet
ws_two = wb.create_sheet(title='my_sheet')
# 20行
for i in range(1, 21):
    # 600列
    # ws_two.append(range(601))
    """
        cell
            row：        单元格所在的行
            column：     单元格坐在的列
            value：      单元格的值
            coordinate： 单元格的坐标
    """
    # 第二列 前20行 都赋值test
    ws_two.cell(column=2, row=i).value = 'test'

# 保存文件
wb.save("sample.xlsx")


