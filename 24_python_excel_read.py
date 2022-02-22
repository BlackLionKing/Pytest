from openpyxl import load_workbook

# 指定文件名
wb = load_workbook(filename='./sample.xlsx')
# 指定文件内的工作表
sheet_ranges = wb['my_sheet']
# 获取文件内工作表的最大行数和列数
row = sheet_ranges.max_row
column = sheet_ranges.max_column

# 循环读取 range从0开始 需手动改为从1开始 row跟column需+1
for i in range(1, row+1):
    for j in range(1, column+1):
        print(sheet_ranges.cell(row=i, column=j).value)
# 读取工作表内a2的值
# print(sheet_ranges['A2'].value)
